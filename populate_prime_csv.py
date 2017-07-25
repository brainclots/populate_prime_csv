#!/usr/bin/env python

'''
Purpose:    Connect to a list of cisco devices, run the commands contained
            in one file (commands) and write the output to a csv file for
            use with Cisco Prime.

Author:
            ___  ____ _ ____ _  _    _  _ _    ____ ___ ___
            |__] |__/ | |__| |\ |    |_/  |    |  |  |    /
            |__] |  \ | |  | | \|    | \_ |___ |__|  |   /__
            Brian.Klotz@nike.com

Version:    0.1
Date:       July 2017
'''

import netmiko
import argparse
import getpass
import openpyxl
import logging

# Set up argument parser and help info
parser = argparse.ArgumentParser(description='Connect to list of devices and '
                                 'obtain specific variable information, '
                                 'writing it to an Excel file that can be '
                                 'exported to a CSV file and imported into Prime.')
always_required = parser.add_argument_group('always required')
always_required.add_argument("template", nargs=1,
                             help="Name of file containing devices and "
                             "variables", metavar='<template>')
args = parser.parse_args()

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.FileHandler('output.log')
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)


def open_file(file):
    wb = openpyxl.load_workbook(file)
    ws1 = wb.active
    ws2 = wb.worksheets[1]
    input_list = []
    command_dict = {}
    for col in range(2, ws1.max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        device = ws1[column_letter + '13'].value
        input_list.append(device)
    for row in range(2, ws2.max_row + 1):
        variable = ws2['A' + str(row)].value
        command = ws2['B' + str(row)].value
        command_dict[variable] = command
    return input_list, command_dict


def get_creds():  # Prompt for credentials
    username = getpass.getuser()
#   username = raw_input('User ID: ')
    try:
        password = getpass.getpass()
        return username, password
    except KeyboardInterrupt:
        print('\n')
        exit()


def obtain_output(connection, command):
    output = connection.send_command(command)
    return output


def main():
    template = args.template[0]
    device_list, command_dict = open_file(template)

    username, password = get_creds()

    netmiko_exceptions = (netmiko.ssh_exception.NetMikoTimeoutException,
                          netmiko.ssh_exception.NetMikoAuthenticationException)

    print('Loaded %d devices' % len(device_list))
    # Build device dictionary
    for a_device in device_list:
        a_device = {'host': a_device,
                    'device_type': 'cisco_ios',
                    'username': username,
                    'password': password,
                    'secret': password
                    }

        # Build command dictionary

        print('-'*80)
        print('Connecting to ' + a_device['host'] + '...')
        try:
            connection = netmiko.ConnectHandler(**a_device)
            connection.enable()

            connection.disconnect()

        except netmiko_exceptions as e:
            print('Failed to connect: %s' % e)

        wb.save(filename=outfile)


main()
